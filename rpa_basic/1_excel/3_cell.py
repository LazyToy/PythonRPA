from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을 땐 'None' 을 출력

# row = 1, 2, 3, ....
# colum = A(1), B(2), C(3), ....
print(ws.cell(row=1, column=1).value) # ws["A1"].value와 똑같은 값을 반환
print(ws.cell(row=1, column=2).value) # ws["B1"].value와 똑같은 값을 반환

c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10 과 같은 값을 반환
print(c.value) # ws["C1"]

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1 
for x in range(1,11): # 10 개 row
    for y in range(1,11): # 10 개의 column
        ws.cell(row=x, column=y, value=randint(0,100)) # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1 


wb.save("sample.xlsx")
